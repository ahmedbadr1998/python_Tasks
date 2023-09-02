
import os
import openpyxl

def read_excel_file(filepath):
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(row)
    
    workbook.close()

def write_excel_file(filepath,mydata):
    workbook=openpyxl.Workbook() 
    sheet=workbook.active
    data=mydata
    for row in data:
        sheet.append(row)    
    workbook.save(filepath)
    workbook.close()  
data=[
    ["Proto type","Unique ID"],
    ]
num=500
with open("/home/ahmed/work/python_workspace/session3_task/header_file.txt","r") as fd:
    
    for line in fd:
        words =line.strip().split('\n')
        words.append('#'+str(num))
        data.append(words)
        num+=1
# print(data)
write_excel_file("/home/ahmed/work/python_workspace/session3_task/tasks/excelheader_file.txt",data)
